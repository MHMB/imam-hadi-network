"""Parse the ``افراد`` (People) sheet.

Layout (SPEC.md §2.2): table ``person`` at B2:N23. Columns of interest:

- B: نام و نام خانوادگی (full name) — primary key in the legacy sheet
- C: شماره تماس (phone) — possibly blank in the sample data
- D: پیامرسان (messenger handle)
- E: رابط/ضامن-4 → role secondary_4
- F: رابط/ضامن-3 → role secondary_3
- G: رابط/ضامن-2 → role secondary_2
- H: رابط/ضامن-اصلی → role main
- I: تایید (verified flag) — "1" / blank

Columns J..N are rolled-up totals computed by Excel formulas; the DB
re-derives them from loan_party + installment so we ignore them.

Person identity in the DB is phone-keyed.  When phone is missing in the
sample, we synthesise a placeholder ``+0__name__`` so the import still
links lender / borrower / guarantor name references.  Real production
data should always carry phones; the importer emits an
``unknown_phone_format`` warning whenever it falls back.
"""

from __future__ import annotations

import unicodedata

from openpyxl.worksheet.worksheet import Worksheet

from app.importer.models import ParsedGuarantorLink, ParsedIssue, ParsedPerson, ParseResult
from app.importer.phone import canonicalize
from app.models.enums import GuarantorRole, IssueCategory, IssueSeverity

_HEADER_ROW = 2  # data starts at row 3 — first data row

# Column indexes (1-based) within the افراد sheet
COL_NAME = 2
COL_PHONE = 3
COL_MESSENGER = 4
COL_GUARANTOR_4 = 5
COL_GUARANTOR_3 = 6
COL_GUARANTOR_2 = 7
COL_GUARANTOR_MAIN = 8
COL_VERIFIED = 9

# Slot column → enum mapping
_GUARANTOR_SLOTS: tuple[tuple[int, GuarantorRole], ...] = (
    (COL_GUARANTOR_MAIN, GuarantorRole.main),
    (COL_GUARANTOR_2, GuarantorRole.secondary_2),
    (COL_GUARANTOR_3, GuarantorRole.secondary_3),
    (COL_GUARANTOR_4, GuarantorRole.secondary_4),
)


def _normalise_name(value: object) -> str:
    """NFC + strip whitespace.  Real production data often has trailing
    spaces or zero-width joiners that break exact-match lookups."""
    if value is None:
        return ""
    return unicodedata.normalize("NFC", str(value)).strip()


def _placeholder_phone(name: str) -> str:
    """Synthetic phone for sample rows where the real one is blank.

    Use a clearly-fake international prefix so it can never collide with
    a real Iranian number and the issues page can pattern-match for it.
    """
    return f"+0__{name}__"


def parse_people(ws: Worksheet, result: ParseResult) -> None:
    """Append every افراد row to ``result.persons`` and link guarantors.

    Emits issues for: malformed/missing phone numbers, ``#REF!`` rolled-up
    formulas (which we ignore for data but flag for visibility), and any
    row whose name fails to parse.
    """
    sheet_name = ws.title
    for r in range(_HEADER_ROW + 1, ws.max_row + 1):
        name = _normalise_name(ws.cell(row=r, column=COL_NAME).value)
        if not name:
            continue

        # --- phone (may be blank in sample data) ---
        phone_cell = ws.cell(row=r, column=COL_PHONE).value
        canonical = canonicalize(phone_cell)
        if canonical.canonical:
            phone_value = canonical.canonical
            phone_raw = canonical.raw or None
        else:
            phone_value = _placeholder_phone(name)
            phone_raw = canonical.raw or None
            result.issues.append(
                ParsedIssue(
                    severity=IssueSeverity.warning,
                    category=IssueCategory.unknown_phone_format,
                    message=(
                        f"شخص «{name}» شماره تماس معتبر ندارد؛ "
                        "از کلید جایگزین برای شناسایی استفاده می‌شود."
                    ),
                    sheet=sheet_name,
                    cell=f"{sheet_name}!C{r}",
                    context={"name": name, "raw": phone_raw},
                )
            )

        # --- guarantor slots ---
        links: list[ParsedGuarantorLink] = []
        for col, role in _GUARANTOR_SLOTS:
            slot_value = ws.cell(row=r, column=col).value
            slot_name = _normalise_name(slot_value)
            if not slot_name:
                continue
            links.append(ParsedGuarantorLink(role=role, guarantor_name=slot_name))

        # --- verified flag ---
        verified_raw = ws.cell(row=r, column=COL_VERIFIED).value
        is_verified = verified_raw is not None and str(verified_raw).strip() == "1"

        # --- messenger ---
        messenger_raw = ws.cell(row=r, column=COL_MESSENGER).value
        messenger = _normalise_name(messenger_raw) or None

        result.persons.append(
            ParsedPerson(
                full_name=name,
                phone_canonical=phone_value,
                phone_raw=phone_raw,
                is_verified=is_verified,
                messenger=messenger,
                guarantor_links=tuple(links),
            )
        )

        # --- broken-formula detection (defensive) ---
        # The Excel rollup columns (J..N) sometimes degrade to #REF! after
        # row inserts/deletes (sample rows 16..23).  Flag for visibility;
        # the DB never reads these values anyway (it recomputes from
        # loan_party + installment), so no behaviour change.
        for col in (10, 11, 12, 13, 14):
            cell = ws.cell(row=r, column=col)
            if isinstance(cell.value, str) and "#REF!" in cell.value:
                result.issues.append(
                    ParsedIssue(
                        severity=IssueSeverity.warning,
                        category=IssueCategory.broken_ref,
                        message=(
                            f"فرمول جمع‌بندی شخص «{name}» در ستون {cell.column_letter}"
                            " به مرجع شکسته شده اشاره می‌کند."
                        ),
                        sheet=sheet_name,
                        cell=f"{sheet_name}!{cell.coordinate}",
                        context={"name": name, "raw": cell.value},
                    )
                )
