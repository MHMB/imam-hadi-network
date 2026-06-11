"""Generic year-sheet parser, driven by :class:`YearLayout` configs.

One walker covers every ledger format the workbook ever used; per-year
differences live entirely in :mod:`app.importer.parsers.layout`.  The
walk itself is the same for all years:

- A loan group starts on any row whose ``ردیف`` (col A) holds a literal
  (non-formula) value; loan-level fields are read from that row.
  Continuation rows inherit the open group.
- Every group/continuation row is one lender contribution: name from
  ``c_lender``, amount from ``c_amount`` (possibly on the row below —
  ``amount_row_offset``), repayment grid decoded per ``grid_encoding``.
- A green fill on an amount cell marks that installment paid
  (see :mod:`app.importer.colors`).

"All data lands" fallbacks (the workbook is the source of truth; the
importer adapts, never the file):

- blank loan number → synthesised ``بدون‌شماره-r<row>`` + warning;
- blank/formula borrower → loan imported without a borrower party + warning;
- blank/zero total → falls back to the sum of lender amounts; if that is
  still not positive the loan cannot satisfy the DB ``total_amount > 0``
  check and is skipped with an *error* issue (observed only on a handful of
  never-funded 1401/1402 rows).
"""

from __future__ import annotations

import unicodedata
from decimal import Decimal, InvalidOperation
from typing import cast

from openpyxl.cell import Cell
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from app.importer.colors import fill_signature, is_green
from app.importer.models import (
    ParsedInstallment,
    ParsedIssue,
    ParsedLoan,
    ParsedParty,
    ParseResult,
)
from app.importer.parsers.layout import DEFAULT_TOPIC, GridEncoding, YearLayout
from app.models.enums import InstallmentStatus, IssueCategory, IssueSeverity, LoanPartyRole

ROW_INDEX_COL = 1  # ردیف — group-boundary marker on every layout


def _normalise_text(value: object) -> str:
    if value is None:
        return ""
    return unicodedata.normalize("NFC", str(value)).strip()


def _is_formula(value: object) -> bool:
    """openpyxl returns formulas as plain strings beginning with ``=``."""
    return isinstance(value, str) and value.startswith("=")


def _literal_text(ws: Worksheet, row: int, col: int | None) -> str | None:
    """Text content of a cell, ignoring formula strings (continuation rows)."""
    if col is None:
        return None
    value = ws.cell(row=row, column=col).value
    if value is None or _is_formula(value):
        return None
    return _normalise_text(value) or None


def _to_decimal(
    value: object,
    *,
    sheet: str,
    cell: str,
    issues: list[ParsedIssue],
) -> Decimal | None:
    if value is None or value == "" or _is_formula(value):
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError):
        issues.append(
            ParsedIssue(
                severity=IssueSeverity.warning,
                category=IssueCategory.bad_day,  # closest category for "unparseable number"
                message=f"مقدار نامعتبر در سلول {cell}: نمی‌توان به عدد تبدیل کرد ({value!r}).",
                sheet=sheet,
                cell=cell,
                context={"raw": str(value)},
            )
        )
        return None


def _to_day_of_month(value: object) -> int | None:
    if value is None or value == "" or _is_formula(value):
        return None
    try:
        return int(float(str(value)))
    except (ValueError, TypeError):
        return None


def _starts_new_group(ws: Worksheet, row: int) -> bool:
    value = ws.cell(row=row, column=ROW_INDEX_COL).value
    if value is None or value == "":
        return False
    return not _is_formula(value)


def _classify(amount_cell: Cell, sheet: str) -> tuple[InstallmentStatus, ParsedIssue | None]:
    """Green fill = paid; any other fill is flagged but treated as unpaid."""
    fill = cast(PatternFill, amount_cell.fill)
    if is_green(fill):
        return InstallmentStatus.paid, None
    sig = fill_signature(fill)
    if sig is None:
        return InstallmentStatus.unpaid, None
    issue = ParsedIssue(
        severity=IssueSeverity.info,
        category=IssueCategory.color_anomaly,
        message=(
            f"رنگ غیرمنتظره روی سلول {amount_cell.coordinate}؛ "
            f"امضای رنگ: {sig}. به عنوان «پرداخت‌نشده» در نظر گرفته شد."
        ),
        sheet=sheet,
        cell=f"{sheet}!{amount_cell.coordinate}",
        context={"signature": sig},
    )
    return InstallmentStatus.unpaid, issue


# ------------------------------------------------------------------ grid


def _grid_cells(
    ws: Worksheet,
    row: int,
    layout: YearLayout,
    pair_index: int,
) -> tuple[Cell | None, Cell]:
    """(day_cell | None, amount_cell) for one month slot of one contribution."""
    enc = layout.grid_encoding
    if enc is GridEncoding.table_pairs:
        day_col = layout.grid_first_col + pair_index * 2
        return (
            cast(Cell, ws.cell(row=row, column=day_col)),
            cast(Cell, ws.cell(row=row, column=day_col + 1)),
        )
    col = layout.grid_first_col + pair_index
    if enc is GridEncoding.paired_day_amount:
        return (
            cast(Cell, ws.cell(row=row, column=col)),
            cast(Cell, ws.cell(row=row + 1, column=col)),
        )
    # paired_amount_only — the sheet records no due day at all.
    return None, cast(Cell, ws.cell(row=row + 1, column=col))


def _read_installments(
    ws: Worksheet,
    row: int,
    layout: YearLayout,
    sheet: str,
    issues: list[ParsedIssue],
) -> list[ParsedInstallment]:
    out: list[ParsedInstallment] = []
    dayless = layout.grid_encoding is GridEncoding.paired_amount_only
    for pair_index, (year, month) in enumerate(layout.grid_months):
        day_cell, amount_cell = _grid_cells(ws, row, layout, pair_index)
        day_val = day_cell.value if day_cell is not None else None
        amount_val = amount_cell.value
        if day_val in (None, "") and amount_val in (None, ""):
            continue
        day = _to_day_of_month(day_val)
        amount = _to_decimal(
            amount_val,
            sheet=sheet,
            cell=f"{sheet}!{amount_cell.coordinate}",
            issues=issues,
        )
        if day is None and amount is None:
            continue
        if day is None:
            # On day-bearing layouts a lone amount is a data gap worth
            # flagging; on the 1403 amount-only layout day 1 is simply the
            # encoding's resolution — not an issue.
            if not dayless and day_cell is not None:
                issues.append(
                    ParsedIssue(
                        severity=IssueSeverity.info,
                        category=IssueCategory.missing_day,
                        message=f"قسط در سلول {amount_cell.coordinate} روز سررسید ندارد.",
                        sheet=sheet,
                        cell=f"{sheet}!{day_cell.coordinate}",
                    )
                )
            day = 1
        if amount is None:
            issues.append(
                ParsedIssue(
                    severity=IssueSeverity.info,
                    category=IssueCategory.missing_amount,
                    message=f"روز سررسید بدون مبلغ در ستون {amount_cell.column_letter}.",
                    sheet=sheet,
                    cell=f"{sheet}!{amount_cell.coordinate}",
                )
            )
            continue
        if amount <= 0:
            # Zero/negative amounts cannot land (DB CHECK amount > 0) and a
            # zero schedule cell carries no information anyway.
            issues.append(
                ParsedIssue(
                    severity=IssueSeverity.info,
                    category=IssueCategory.missing_amount,
                    message=f"قسط با مبلغ صفر در سلول {amount_cell.coordinate} نادیده گرفته شد.",
                    sheet=sheet,
                    cell=f"{sheet}!{amount_cell.coordinate}",
                )
            )
            continue
        status, color_issue = _classify(amount_cell, sheet)
        if color_issue is not None:
            issues.append(color_issue)
        out.append(
            ParsedInstallment(
                due_persian_year=year,
                due_persian_month=month,
                due_day_of_month=day,
                amount=amount,
                status=status,
                sheet=sheet,
                cell=f"{sheet}!{amount_cell.coordinate}",
            )
        )
    return out


# ------------------------------------------------------------------ walker


class _OpenLoan:
    """Mutable accumulator for the loan group currently being walked."""

    __slots__ = (
        "borrower",
        "channel",
        "description",
        "guarantor",
        "liaison",
        "loan_number",
        "parties",
        "row",
        "topic",
        "total",
    )

    def __init__(self) -> None:
        self.row = 0
        self.loan_number: str | None = None
        self.borrower: str | None = None
        self.total: Decimal | None = None
        self.topic: str | None = None
        self.guarantor: str | None = None
        self.channel: str | None = None
        self.liaison: str | None = None
        self.description: str | None = None
        self.parties: list[ParsedParty] = []


def _open_group(
    ws: Worksheet,
    row: int,
    layout: YearLayout,
    sheet: str,
    issues: list[ParsedIssue],
) -> _OpenLoan:
    """Read the loan-level fields from a group-start row."""
    loan = _OpenLoan()
    loan.row = row
    loan.loan_number = _literal_text(ws, row, layout.c_loan_number)
    loan.borrower = _literal_text(ws, row, layout.c_borrower)
    loan.topic = _literal_text(ws, row, layout.c_topic)
    loan.guarantor = _literal_text(ws, row, layout.c_guarantor)
    loan.liaison = _literal_text(ws, row, layout.c_liaison)
    loan.description = _literal_text(ws, row, layout.c_description)
    # The legacy sheets store absent channel numbers as literal 0.
    channel = _literal_text(ws, row, layout.c_channel)
    loan.channel = None if channel == "0" else channel
    loan.total = _to_decimal(
        ws.cell(row=row, column=layout.c_total).value,
        sheet=sheet,
        cell=f"{sheet}!{ws.cell(row=row, column=layout.c_total).coordinate}",
        issues=issues,
    )
    return loan


def parse_year(
    ws: Worksheet,
    persian_year: int,
    layout: YearLayout,
    result: ParseResult,
) -> None:
    """Append every loan of one year sheet to ``result.loans``."""
    sheet = ws.title
    open_loan: _OpenLoan | None = None

    def _flush() -> None:
        nonlocal open_loan
        loan, open_loan = open_loan, None
        if loan is None:
            return

        loan_number = loan.loan_number
        if loan_number is None:
            loan_number = f"بدون‌شماره-r{loan.row}"
            result.issues.append(
                ParsedIssue(
                    severity=IssueSeverity.warning,
                    category=IssueCategory.orphan_row,
                    message=(
                        f"گروه قرض در ردیف {loan.row} شماره (ش) ندارد؛ "
                        f"شناسه «{loan_number}» جایگزین شد."
                    ),
                    sheet=sheet,
                    cell=f"{sheet}!B{loan.row}",
                )
            )

        total = loan.total
        if total is None or total <= 0:
            lender_sum = sum((p.amount for p in loan.parties), Decimal(0))
            if lender_sum > 0:
                result.issues.append(
                    ParsedIssue(
                        severity=IssueSeverity.warning,
                        category=IssueCategory.total_mismatch,
                        message=(
                            f"قرض #{loan_number}: مبلغ کل خالی/صفر است؛ "
                            f"مجموع قرض‌دهندگان ({lender_sum}) جایگزین شد."
                        ),
                        sheet=sheet,
                        cell=f"{sheet}!{loan.row}",
                        context={"loan_number": loan_number, "lender_sum": str(lender_sum)},
                    )
                )
                total = lender_sum
            else:
                result.issues.append(
                    ParsedIssue(
                        severity=IssueSeverity.error,
                        category=IssueCategory.orphan_row,
                        message=(
                            f"قرض #{loan_number} (ردیف {loan.row}) نه مبلغ کل دارد و نه "
                            "قرض‌دهنده‌ای؛ قابل ثبت نیست و نادیده گرفته شد."
                        ),
                        sheet=sheet,
                        cell=f"{sheet}!{loan.row}",
                        context={"loan_number": loan_number, "row": loan.row},
                    )
                )
                return

        parties: list[ParsedParty] = []
        if loan.borrower:
            parties.append(
                ParsedParty(
                    role=LoanPartyRole.borrower,
                    person_name=loan.borrower,
                    amount=total,
                    display_order=0,
                )
            )
        else:
            result.issues.append(
                ParsedIssue(
                    severity=IssueSeverity.warning,
                    category=IssueCategory.unresolved_person,
                    message=f"قرض #{loan_number}: قرض‌گیرنده نامشخص است.",
                    sheet=sheet,
                    cell=f"{sheet}!{loan.row}",
                    context={"loan_number": loan_number},
                )
            )
        parties.extend(loan.parties)

        topic = loan.topic
        if topic is None:
            if layout.c_topic is not None:
                result.issues.append(
                    ParsedIssue(
                        severity=IssueSeverity.warning,
                        category=IssueCategory.unknown_topic,
                        message=(
                            f"قرض #{loan_number}: موضوع خالی است؛ "
                            f"«{DEFAULT_TOPIC}» در نظر گرفته شد."
                        ),
                        sheet=sheet,
                        cell=f"{sheet}!{loan.row}",
                        context={"loan_number": loan_number},
                    )
                )
            topic = DEFAULT_TOPIC

        result.loans.append(
            ParsedLoan(
                persian_year=persian_year,
                loan_number=loan_number,
                total_amount=total,
                topic_name=topic,
                parties=tuple(parties),
                channel_number=loan.channel,
                guarantor_name=loan.guarantor,
                liaison_label=loan.liaison,
                description=loan.description,
                source_sheet=sheet,
            )
        )

    row = layout.first_data_row
    while row <= ws.max_row:
        if _starts_new_group(ws, row):
            _flush()
            open_loan = _open_group(ws, row, layout, sheet, result.issues)

        lender = _literal_text(ws, row, layout.c_lender)
        amount_cell = ws.cell(row=row + layout.amount_row_offset, column=layout.c_amount)
        amount = _to_decimal(
            amount_cell.value,
            sheet=sheet,
            cell=f"{sheet}!{amount_cell.coordinate}",
            issues=result.issues,
        )

        if (lender is not None or amount is not None) and open_loan is not None:
            installments = _read_installments(ws, row, layout, sheet, result.issues)
            open_loan.parties.append(
                ParsedParty(
                    role=LoanPartyRole.lender,
                    person_name=lender or "",
                    amount=amount if amount is not None else Decimal(0),
                    display_order=len(open_loan.parties),
                    installments=tuple(installments),
                )
            )
            if lender is None:
                result.issues.append(
                    ParsedIssue(
                        severity=IssueSeverity.warning,
                        category=IssueCategory.unresolved_person,
                        message=f"وام‌دهنده در ردیف {row} نامشخص است.",
                        sheet=sheet,
                        cell=f"{sheet}!{ws.cell(row=row, column=layout.c_lender).coordinate}",
                    )
                )

        row += layout.row_step

    _flush()
