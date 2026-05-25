"""Cell-fill colour detection for the xlsm importer.

The legacy workbook signals "this installment is paid" with a green cell
fill (``#00B050``).  The VBA helper ``SumifColor`` (see SPEC.md §2.1 and
the ``مانده`` formulas in §2.6) sums these greens as the paid portion.

We detect the same colour at parse time.  ``openpyxl`` exposes fills in
a few different shapes — explicit RGB, theme-tinted, or indexed —
``is_green`` handles all of them but only matches the **literal**
``00B050`` family.  Theme-derived greens that *look* the same to a human
are reported by the importer as ``color_anomaly`` warnings (Phase 2 can
extend the matcher if needed).
"""

from __future__ import annotations

from openpyxl.styles import Color, PatternFill

# Canonical green used by the workbook.  Last 6 hex chars (RGB) compared
# case-insensitively; the leading byte is alpha, which we ignore.
_GREEN_RGB = "00B050"


def is_green(fill: PatternFill | None) -> bool:
    """Return ``True`` if a cell's fill is the workbook's "paid" green.

    Accepts ``None`` and unstyled fills (returns ``False``) so callers
    don't have to guard every access.
    """
    if fill is None or not getattr(fill, "patternType", None):
        return False
    fg: Color | None = getattr(fill, "fgColor", None)
    if fg is None:
        return False

    # openpyxl returns sentinel objects for un-applicable attributes
    # (`Values must be of type <class 'str'>` etc).  Wrap in try/except
    # rather than touching the private API.
    try:
        if getattr(fg, "type", None) == "rgb":
            rgb = fg.rgb
            if isinstance(rgb, str) and rgb.upper().endswith(_GREEN_RGB):
                return True
    except (TypeError, AttributeError):
        pass

    return False


def fill_signature(fill: PatternFill | None) -> str | None:
    """Stable string label for a fill — used in ``color_anomaly`` issue context.

    Returns ``None`` for unstyled cells.  Output is intentionally short
    and uppercase so admins can pattern-match in the issues page.
    """
    if fill is None or not getattr(fill, "patternType", None):
        return None
    fg: Color | None = getattr(fill, "fgColor", None)
    if fg is None:
        return None

    fg_type = getattr(fg, "type", None)
    try:
        if fg_type == "rgb" and isinstance(fg.rgb, str):
            label = f"rgb:{fg.rgb.upper()}"
        elif fg_type == "theme":
            label = f"theme:{fg.theme}/tint:{round(float(fg.tint), 2)}"
        elif fg_type == "indexed":
            label = f"indexed:{fg.indexed}"
        else:
            label = "unknown"
    except (TypeError, AttributeError, ValueError):
        label = "unknown"
    return label
