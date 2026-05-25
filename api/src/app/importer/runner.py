"""End-to-end import orchestration.

Wraps the parse → validate → write pipeline so both the CLI (P2.9) and
the HTTP background task (P6) call the same code.  No I/O beyond reading
the xlsm and writing to the DB.
"""

from __future__ import annotations

import time
from collections.abc import AsyncIterator
from contextlib import asynccontextmanager
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from app.config import settings
from app.importer.models import ParseResult
from app.importer.parsers import detect_year_sheets, year_parser_for
from app.importer.parsers.people import parse_people
from app.importer.parsers.topics import parse_topics
from app.importer.validation import validate
from app.importer.writer import sha256_of_file, write_parse_result
from app.logging import get_logger
from app.models import Import
from app.models.enums import ImportStatus

log = get_logger(__name__)


@dataclass(slots=True)
class ImportOutcome:
    """Summary of one import run — what the CLI prints and the API returns."""

    file: Path
    sha256: str
    import_id: int
    years_imported: list[int]
    loans: int
    persons: int
    topics_in_file: int
    issues_total: int
    issues_by_severity: dict[str, int]
    issues_by_category: dict[str, int]
    error_count: int
    deduped: bool
    duration_ms: int


def parse_workbook(path: Path) -> ParseResult:
    """Open the xlsm and run every sub-parser; return populated ParseResult.

    Pure: no DB, no validation — see ``run_import`` for the full pipeline.
    """
    wb = openpyxl.load_workbook(path, data_only=False, keep_vba=True)
    result = ParseResult()

    if "موضوعات" in wb.sheetnames:
        parse_topics(wb["موضوعات"], result)
    if "افراد" in wb.sheetnames:
        parse_people(wb["افراد"], result)

    for year, sheet_name in detect_year_sheets(wb):
        result.years_present.append(year)
        parser = year_parser_for(year)
        parser(wb[sheet_name], year, result)

    return result


def _build_outcome(
    *,
    path: Path,
    sha: str,
    import_row: Import,
    result: ParseResult,
    duration_ms: int,
    deduped: bool,
) -> ImportOutcome:
    by_sev: dict[str, int] = {}
    by_cat: dict[str, int] = {}
    for issue in result.issues:
        by_sev[issue.severity.value] = by_sev.get(issue.severity.value, 0) + 1
        by_cat[issue.category.value] = by_cat.get(issue.category.value, 0) + 1
    return ImportOutcome(
        file=path,
        sha256=sha,
        import_id=import_row.id,
        years_imported=list(import_row.years_imported),
        loans=len(result.loans),
        persons=len(result.persons),
        topics_in_file=len(result.topics),
        issues_total=len(result.issues),
        issues_by_severity=by_sev,
        issues_by_category=by_cat,
        error_count=by_sev.get("error", 0),
        deduped=deduped,
        duration_ms=duration_ms,
    )


async def run_import(
    session: AsyncSession,
    path: Path,
    *,
    dry_run: bool = False,
) -> ImportOutcome:
    """End-to-end pipeline: hash → parse → validate → (write or skip).

    ``dry_run=True`` runs parse + validate against the file and returns
    a synthetic outcome describing what *would* land, with no DB writes.
    """
    started = time.monotonic()
    sha = sha256_of_file(path)
    log.info("importer.start", file=str(path), sha256=sha, dry_run=dry_run)

    result = parse_workbook(path)
    validate(result)
    duration_ms = int((time.monotonic() - started) * 1000)

    if dry_run:
        # Synthetic Import row, not persisted.
        synthetic = Import(
            source_sha256=sha,
            source_filename=path.name,
            years_imported=sorted({loan.persian_year for loan in result.loans}),
            status=ImportStatus.success,
            duration_ms=duration_ms,
            report={},
        )
        # The synthetic Import isn't persisted; give it a sentinel id for the
        # outcome dataclass.  -1 is fine because real ids are positive serials.
        synthetic.id = -1
        return _build_outcome(
            path=path,
            sha=sha,
            import_row=synthetic,
            result=result,
            duration_ms=duration_ms,
            deduped=False,
        )

    pre_write = time.monotonic()
    import_row = await write_parse_result(
        session,
        source_path=path,
        sha256=sha,
        duration_ms=duration_ms,
        result=result,
    )
    write_ms = int((time.monotonic() - pre_write) * 1000)
    # A write that took less than this and whose persisted duration_ms differs
    # from the one we just measured must be a sha-dedup short-circuit (we
    # returned an existing Import row, no INSERTs ran).
    _DEDUP_WRITE_THRESHOLD_MS = 50
    deduped = write_ms < _DEDUP_WRITE_THRESHOLD_MS and import_row.duration_ms != duration_ms
    total_ms = int((time.monotonic() - started) * 1000)

    log.info(
        "importer.done",
        file=str(path),
        import_id=import_row.id,
        years_imported=list(import_row.years_imported),
        loans=len(result.loans),
        issues=len(result.issues),
        deduped=deduped,
        duration_ms=total_ms,
    )
    return _build_outcome(
        path=path,
        sha=sha,
        import_row=import_row,
        result=result,
        duration_ms=total_ms,
        deduped=deduped,
    )


@asynccontextmanager
async def open_session() -> AsyncIterator[AsyncSession]:
    """Convenience: build a fresh async session bound to the configured DB."""
    engine = create_async_engine(str(settings.database_url))
    sessionmaker = async_sessionmaker(engine, expire_on_commit=False)
    try:
        async with sessionmaker() as session:
            yield session
    finally:
        await engine.dispose()
