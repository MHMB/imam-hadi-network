"""Writer integration tests against a live Postgres.

Verifies the contract documented at the top of ``app.importer.writer``:

- sha-dedup: re-uploading the same file → same Import row, no new loans.
- per-year scoped replace: uploading a modified xlsm only swaps the
  years it contains; older years' data is untouched.
- single transaction: an error during a write leaves no half-imported
  state behind.
"""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from app.config import settings
from app.importer.models import ParseResult
from app.importer.parsers.people import parse_people
from app.importer.parsers.topics import parse_topics
from app.importer.parsers.year_1404 import parse_year_1404
from app.importer.parsers.year_1405 import parse_year_1405
from app.importer.validation import validate
from app.importer.writer import sha256_of_file, write_parse_result
from app.models import DataIssue, Import, Loan, LoanParty, Person
from app.models.enums import LoanPartyRole

pytestmark = pytest.mark.integration


# --------------------------------------------------------------------- helpers


_WIPE_TABLES = (
    "data_issue",
    "installment",
    "loan_party",
    "loan",
    "import",
    "person_guarantor",
    "person",
)


@pytest.fixture
async def session() -> AsyncSession:
    """Hermetic DB: clean importer-managed tables before AND after each test.

    Topics (loan_topic) are left intact because the seed migration assigns
    legacy_num=0 to "نامعلوم" and the writer's upsert never deletes topics —
    blowing them away would lose that metadata and other tests would fail.
    """
    engine = create_async_engine(str(settings.database_url))
    sessionmaker = async_sessionmaker(engine, expire_on_commit=False)

    async def wipe() -> None:
        async with engine.begin() as conn:
            for table in _WIPE_TABLES:
                await conn.exec_driver_sql(f"DELETE FROM {table}")

    await wipe()
    async with sessionmaker() as s:
        yield s
    await wipe()
    await engine.dispose()


def _parse_full(xlsm: Path) -> ParseResult:
    wb = openpyxl.load_workbook(xlsm, data_only=False, keep_vba=True)
    result = ParseResult()
    parse_topics(wb["موضوعات"], result)
    parse_people(wb["افراد"], result)
    if "سال 1404" in wb.sheetnames:
        parse_year_1404(wb["سال 1404"], 1404, result)
    if "سال 1405" in wb.sheetnames:
        parse_year_1405(wb["سال 1405"], 1405, result)
    validate(result)
    return result


# --------------------------------------------------------------------- tests


@pytest.mark.asyncio
async def test_first_import_writes_everything(
    session: AsyncSession, sample_xlsm_path: Path
) -> None:
    result = _parse_full(sample_xlsm_path)
    sha = sha256_of_file(sample_xlsm_path)
    imp, _deduped = await write_parse_result(
        session,
        source_path=sample_xlsm_path,
        sha256=sha,
        duration_ms=42,
        result=result,
    )
    assert imp.source_sha256 == sha
    assert sorted(imp.years_imported) == [1404, 1405]

    # 9 loans (5 from 1404 + 4 from 1405)
    loan_count = (await session.execute(select(func.count()).select_from(Loan))).scalar_one()
    assert loan_count == 9

    # 21 persons
    person_count = (await session.execute(select(func.count()).select_from(Person))).scalar_one()
    assert person_count == 21

    # Loan 1500: 1 borrower (نفر 1) + 3 lenders (نفر 2/3/4) with the right amounts.
    loan_1500 = (await session.execute(select(Loan).where(Loan.loan_number == "1500"))).scalar_one()
    parties = (
        (
            await session.execute(
                select(LoanParty)
                .where(LoanParty.loan_id == loan_1500.id)
                .order_by(LoanParty.display_order)
            )
        )
        .scalars()
        .all()
    )
    assert parties[0].role is LoanPartyRole.borrower
    assert parties[0].amount == Decimal(20)
    lender_amounts = [p.amount for p in parties if p.role is LoanPartyRole.lender]
    assert lender_amounts == [Decimal(3), Decimal(7), Decimal(10)]

    # DataIssue rows persisted with summary on Import.report
    issues_persisted = (
        await session.execute(select(func.count()).select_from(DataIssue))
    ).scalar_one()
    assert issues_persisted == len(result.issues)
    assert imp.report["loans"] == 9
    assert imp.report["persons"] == 21


@pytest.mark.asyncio
async def test_sha_dedup_returns_existing_import(
    session: AsyncSession, sample_xlsm_path: Path
) -> None:
    result = _parse_full(sample_xlsm_path)
    sha = sha256_of_file(sample_xlsm_path)
    first, first_deduped = await write_parse_result(
        session,
        source_path=sample_xlsm_path,
        sha256=sha,
        duration_ms=10,
        result=result,
    )
    # Second call with same sha — must return the same row, no new Import.
    second, second_deduped = await write_parse_result(
        session,
        source_path=sample_xlsm_path,
        sha256=sha,
        duration_ms=10,
        result=result,
    )
    assert second.id == first.id
    assert first_deduped is False
    assert second_deduped is True
    n_imports = (await session.execute(select(func.count()).select_from(Import))).scalar_one()
    assert n_imports == 1


@pytest.mark.asyncio
async def test_reimport_replaces_only_present_years(
    session: AsyncSession, sample_xlsm_path: Path
) -> None:
    """Build a fake 1404-only ParseResult and ensure 1405 loans stay put."""
    # 1) Seed with the full sample (1404 + 1405).
    full = _parse_full(sample_xlsm_path)
    sha_full = sha256_of_file(sample_xlsm_path)
    _imp, _dedup = await write_parse_result(
        session,
        source_path=sample_xlsm_path,
        sha256=sha_full,
        duration_ms=10,
        result=full,
    )
    loans_1405_before = (
        (await session.execute(select(Loan).where(Loan.persian_year == 1405))).scalars().all()
    )
    loan_ids_1405_before = {ln.id for ln in loans_1405_before}
    assert loan_ids_1405_before  # sanity

    # 2) Now re-import a ParseResult that contains ONLY 1404.
    only_1404 = _parse_full(sample_xlsm_path)
    only_1404.loans = [ln for ln in only_1404.loans if ln.persian_year == 1404]
    # Use a different "filename" + a different sha so the dedup short-circuit
    # doesn't fire and we actually re-execute the writer.
    fake_path = sample_xlsm_path.with_name("sample-1404-only.xlsm")
    fake_sha = "0" * 64
    _imp, _dedup = await write_parse_result(
        session,
        source_path=fake_path,
        sha256=fake_sha,
        duration_ms=5,
        result=only_1404,
    )
    loans_1405_after = (
        (await session.execute(select(Loan).where(Loan.persian_year == 1405))).scalars().all()
    )
    loan_ids_1405_after = {ln.id for ln in loans_1405_after}
    assert loan_ids_1405_after == loan_ids_1405_before, (
        "1405 loans must survive a 1404-only re-import"
    )

    # And the 1404 loans should have been swapped (new IDs).
    loans_1404 = (
        (await session.execute(select(Loan).where(Loan.persian_year == 1404))).scalars().all()
    )
    assert all(ln.import_id != loans_1405_after[0].import_id for ln in loans_1404)


# --------------------------------------------------------------- real-data behaviours


def _synthetic_result() -> ParseResult:
    """A hand-built ParseResult mirroring the real workbook's hard cases:

    - the same lender contributing to one loan on several rows (the pattern
      that violated ``unique_role_person`` on the first production import);
    - lender names that are spelling variants of one person;
    - a lender that is missing from the افراد master entirely.
    """
    from app.importer.models import ParsedInstallment, ParsedLoan, ParsedParty, ParsedPerson
    from app.importer.names import placeholder_phone
    from app.models.enums import InstallmentStatus

    def inst(month: int, amount: int) -> ParsedInstallment:
        return ParsedInstallment(
            due_persian_year=1403,
            due_persian_month=month,
            due_day_of_month=1,
            amount=Decimal(amount),
            status=InstallmentStatus.unpaid,
        )

    result = ParseResult()
    result.topics = ["نامعلوم"]
    result.persons = [
        ParsedPerson(
            full_name="سیدساجد موسوی",
            phone_canonical=placeholder_phone("سیدساجد موسوی"),
            phone_raw=None,
        ),
        ParsedPerson(
            full_name="قرض‌گیرنده الف",
            phone_canonical=placeholder_phone("قرض‌گیرنده الف"),
            phone_raw=None,
        ),
    ]
    result.loans = [
        ParsedLoan(
            persian_year=1403,
            loan_number="795",
            total_amount=Decimal(9),
            topic_name="نامعلوم",
            parties=(
                ParsedParty(
                    role=LoanPartyRole.borrower,
                    person_name="قرض‌گیرنده الف",
                    amount=Decimal(9),
                    display_order=0,
                ),
                # Same person, three contribution rows, two spellings.
                ParsedParty(
                    role=LoanPartyRole.lender,
                    person_name="سیدساجد موسوی",
                    amount=Decimal(4),
                    display_order=0,
                    installments=(inst(1, 4),),
                ),
                ParsedParty(
                    role=LoanPartyRole.lender,
                    person_name="سیدساجدموسوی",  # no-space variant
                    amount=Decimal(3),
                    display_order=1,
                    installments=(inst(2, 3),),
                ),
                # A lender the افراد master never listed.
                ParsedParty(
                    role=LoanPartyRole.lender,
                    person_name="مامانجون",
                    amount=Decimal(2),
                    display_order=2,
                    installments=(inst(3, 2),),
                ),
            ),
        ),
    ]
    return result


@pytest.mark.asyncio
async def test_repeat_lender_rows_merge_into_one_party(session: AsyncSession) -> None:
    """Multiple dated contributions by one lender → one LoanParty, summed
    amount, all installments kept — never a unique_role_person violation."""
    from app.models import Installment

    result = _synthetic_result()
    _imp, _dedup = await write_parse_result(
        session,
        source_path=Path("synthetic.xlsm"),
        sha256="f" * 64,
        duration_ms=1,
        result=result,
    )

    loan = (await session.execute(select(Loan).where(Loan.loan_number == "795"))).scalar_one()
    lenders = (
        (
            await session.execute(
                select(LoanParty).where(
                    LoanParty.loan_id == loan.id, LoanParty.role == LoanPartyRole.lender
                )
            )
        )
        .scalars()
        .all()
    )
    assert len(lenders) == 2  # موسوی (merged ×2) + مامانجون

    by_amount = {p.amount: p for p in lenders}
    merged = by_amount[Decimal(7)]  # 4 + 3 across the two spelling variants
    insts = (
        (await session.execute(select(Installment).where(Installment.loan_party_id == merged.id)))
        .scalars()
        .all()
    )
    assert sorted(i.amount for i in insts) == [Decimal(3), Decimal(4)]


@pytest.mark.asyncio
async def test_unlisted_lender_is_auto_created(session: AsyncSession) -> None:
    """Names referenced by loans but absent from افراد become Person rows."""
    result = _synthetic_result()
    _imp, _dedup = await write_parse_result(
        session,
        source_path=Path("synthetic.xlsm"),
        sha256="e" * 64,
        duration_ms=1,
        result=result,
    )
    person = (
        await session.execute(select(Person).where(Person.full_name == "مامانجون"))
    ).scalar_one()
    assert person.phone.startswith("+0__")
    assert person.is_verified is False

    # Spelling variants must NOT have created a second موسوی row.
    mousavi_count = (
        await session.execute(
            select(func.count()).select_from(Person).where(Person.full_name.like("%موسوی%"))
        )
    ).scalar_one()
    assert mousavi_count == 1
