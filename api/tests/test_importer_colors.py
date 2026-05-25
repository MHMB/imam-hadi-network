"""Tests for ``app.importer.colors`` — cell-fill green detection."""

from __future__ import annotations

from pathlib import Path

from openpyxl.styles import Color, PatternFill

from app.importer.colors import fill_signature, is_green


def _green_fill() -> PatternFill:
    return PatternFill(patternType="solid", fgColor=Color(rgb="FF00B050"))


def _other_rgb_fill(rgb: str) -> PatternFill:
    return PatternFill(patternType="solid", fgColor=Color(rgb=rgb))


def _theme_fill() -> PatternFill:
    return PatternFill(patternType="solid", fgColor=Color(theme=9, tint=0.6))


def _no_fill() -> PatternFill:
    return PatternFill()


def test_green_rgb_detected() -> None:
    assert is_green(_green_fill()) is True


def test_green_rgb_case_insensitive() -> None:
    assert is_green(_other_rgb_fill("ff00b050")) is True


def test_green_without_alpha_byte() -> None:
    # ``Color`` always normalises to 8 chars, but be defensive against
    # synthetic inputs (e.g. user-edited xlsm with bare 6-char RGB).
    cell = PatternFill(patternType="solid", fgColor=Color(rgb="0000B050"))
    assert is_green(cell) is True


def test_non_green_rgb_rejected() -> None:
    assert is_green(_other_rgb_fill("FF0070C0")) is False  # blue header
    assert is_green(_other_rgb_fill("FFBDD6EE")) is False  # pale blue
    assert is_green(_other_rgb_fill("FF00B0F0")) is False  # cyan


def test_theme_green_not_matched() -> None:
    # Theme-derived colours are intentionally rejected — only the literal
    # 00B050 is the workbook's "paid" sentinel.  Anything else is reported
    # as a color_anomaly by the importer.
    assert is_green(_theme_fill()) is False


def test_empty_or_none_safe() -> None:
    assert is_green(None) is False
    assert is_green(_no_fill()) is False


def test_signature_shapes() -> None:
    assert fill_signature(_green_fill()) == "rgb:FF00B050"
    assert fill_signature(_other_rgb_fill("ffabcdef")) == "rgb:FFABCDEF"
    assert fill_signature(_theme_fill()) == "theme:9/tint:0.6"
    assert fill_signature(None) is None
    assert fill_signature(_no_fill()) is None


def test_sample_xlsm_real_cells(sample_xlsm_path: Path) -> None:
    """Cross-check the helper against actual cells in the sample workbook."""
    import openpyxl

    wb = openpyxl.load_workbook(sample_xlsm_path, data_only=False, keep_vba=True)
    ws = wb["سال 1404"]

    # Per SPEC.md §2.6, loan 1500 (row 4-5) has the lender نفر 2 fully paid.
    # The amount row is r=5; the column for شهریور04 ("Shahrivar 1404") is U=21.
    paid_amount_cell = ws.cell(row=5, column=21)
    assert paid_amount_cell.value == 3
    assert is_green(paid_amount_cell.fill) is True

    # An empty cell on the same row is not green.
    blank_cell = ws.cell(row=5, column=22)
    assert blank_cell.value is None
    assert is_green(blank_cell.fill) is False
