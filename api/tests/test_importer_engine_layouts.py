"""Engine tests for the pre-1404 layouts (1401/1402 paired, 1403 amount-only)
and the engine-level "all data lands" fallbacks.

Fixtures are synthesised in-memory with openpyxl — the real workbook is
member PII and never enters the repo.  Geometry mirrors real_data.xlsm
exactly (verified cell-by-cell against the production file).
"""

from __future__ import annotations

from decimal import Decimal

import openpyxl
import pytest
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from app.importer.models import ParseResult
from app.importer.parsers.engine import parse_year
from app.importer.parsers.layout import DEFAULT_TOPIC, LAYOUTS, GridEncoding, layout_for
from app.models.enums import InstallmentStatus, IssueCategory, IssueSeverity, LoanPartyRole

GREEN = PatternFill(fill_type="solid", fgColor="FF00B050")


def _ws(title: str) -> Worksheet:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    return ws


# --------------------------------------------------------------------- layout registry


def test_every_historical_year_has_a_layout() -> None:
    for year in (1401, 1402, 1403, 1404, 1405):
        assert layout_for(year) is not None, year


def test_future_years_rebase_the_table_layout() -> None:
    layout = layout_for(1406)
    assert layout is not None
    assert layout.grid_encoding is GridEncoding.table_pairs
    assert layout.grid_months[0] == (1406, 1)


def test_pre_1401_years_have_no_layout() -> None:
    assert layout_for(1399) is None


def test_month_spans_match_the_real_workbook() -> None:
    assert LAYOUTS[1401].grid_months[0] == (1401, 1)
    assert LAYOUTS[1401].grid_months[-1] == (1402, 12)
    assert LAYOUTS[1402].grid_months[-1] == (1403, 12)
    assert LAYOUTS[1403].grid_months[-1] == (1405, 2)  # 26 months
    assert LAYOUTS[1404].grid_months[-1] == (1406, 2)  # 26 months
    assert LAYOUTS[1405].grid_months[-1] == (1406, 5)  # 17 months


# --------------------------------------------------------------------- 1401 paired layout


def _build_1401_sheet() -> Worksheet:
    """Two loans in the 1401 geometry: B=loan# E=total F=borrower J=guarantor
    K=lender, amount on the bottom row col L, grid P.. day-top/amount-bottom."""
    ws = _ws("سال 1401")
    # Loan 204 — two lenders, second one paid (green) in month 2.
    ws.cell(row=3, column=1, value=1)  # ردیف → group start
    ws.cell(row=3, column=2, value=204)
    ws.cell(row=3, column=3, value=183)  # channel
    ws.cell(row=3, column=4, value="رابط الف")
    ws.cell(row=3, column=5, value=13)  # total
    ws.cell(row=3, column=6, value="قرض‌گیرنده الف")
    ws.cell(row=3, column=8, value="توضیح")
    ws.cell(row=3, column=10, value="ضامن الف")
    ws.cell(row=3, column=11, value="وام‌دهنده یک")
    ws.cell(row=4, column=12, value=10)  # amount (bottom row)
    ws.cell(row=3, column=16, value=15)  # day, month 1 (P top)
    ws.cell(row=4, column=16, value=10)  # amount, month 1 (P bottom)
    # continuation pair → second lender
    ws.cell(row=5, column=11, value="وام‌دهنده دو")
    ws.cell(row=6, column=12, value=3)
    ws.cell(row=5, column=17, value=20)  # day, month 2
    paid = ws.cell(row=6, column=17, value=3)  # amount, month 2
    paid.fill = GREEN
    # Loan 205 — single lender.
    ws.cell(row=7, column=1, value=2)
    ws.cell(row=7, column=2, value=205)
    ws.cell(row=7, column=5, value=5)
    ws.cell(row=7, column=6, value="قرض‌گیرنده ب")
    ws.cell(row=7, column=11, value="وام‌دهنده یک")
    ws.cell(row=8, column=12, value=5)
    return ws


@pytest.fixture
def parsed_1401() -> ParseResult:
    result = ParseResult()
    parse_year(_build_1401_sheet(), 1401, LAYOUTS[1401], result)
    return result


def test_1401_loan_fields(parsed_1401: ParseResult) -> None:
    assert len(parsed_1401.loans) == 2
    loan = parsed_1401.loans[0]
    assert loan.loan_number == "204"
    assert loan.total_amount == Decimal(13)
    assert loan.channel_number == "183"
    assert loan.guarantor_name == "ضامن الف"
    assert loan.liaison_label == "رابط الف"
    assert loan.topic_name == DEFAULT_TOPIC  # no topic column in 1401


def test_1401_parties_and_amount_on_bottom_row(parsed_1401: ParseResult) -> None:
    loan = parsed_1401.loans[0]
    borrower = loan.parties[0]
    assert borrower.role is LoanPartyRole.borrower
    assert borrower.amount == Decimal(13)
    lenders = [p for p in loan.parties if p.role is LoanPartyRole.lender]
    assert [(p.person_name, p.amount) for p in lenders] == [
        ("وام‌دهنده یک", Decimal(10)),
        ("وام‌دهنده دو", Decimal(3)),
    ]


def test_1401_grid_day_amount_rows(parsed_1401: ParseResult) -> None:
    lenders = [p for p in parsed_1401.loans[0].parties if p.role is LoanPartyRole.lender]
    first = lenders[0].installments
    assert len(first) == 1
    assert (first[0].due_persian_year, first[0].due_persian_month) == (1401, 1)
    assert first[0].due_day_of_month == 15
    assert first[0].status is InstallmentStatus.unpaid
    second = lenders[1].installments
    assert (second[0].due_persian_year, second[0].due_persian_month) == (1401, 2)
    assert second[0].status is InstallmentStatus.paid  # green fill


# --------------------------------------------------------------------- 1403 amount-only grid


def _build_1403_sheet() -> Worksheet:
    """1403 geometry: header r3, data r4, E=borrower F=total J=lender K=amount
    (top row), grid M.. amounts on the bottom row, no due-day row."""
    ws = _ws("سال 1403")
    ws.cell(row=4, column=1, value=1)
    ws.cell(row=4, column=2, value=795)
    ws.cell(row=4, column=5, value="قرض‌گیرنده ج")
    ws.cell(row=4, column=6, value=4)  # total
    ws.cell(row=4, column=10, value="صندوق امام هادی")  # lender (alias variant)
    ws.cell(row=4, column=11, value=4)  # amount on the lender's own row
    paid = ws.cell(row=5, column=13, value=2)  # month 1 amount (bottom row)
    paid.fill = GREEN
    ws.cell(row=5, column=15, value=2)  # month 3 amount, unpaid
    return ws


@pytest.fixture
def parsed_1403() -> ParseResult:
    result = ParseResult()
    parse_year(_build_1403_sheet(), 1403, LAYOUTS[1403], result)
    return result


def test_1403_amount_only_grid_defaults_day_silently(parsed_1403: ParseResult) -> None:
    (loan,) = parsed_1403.loans
    (lender,) = [p for p in loan.parties if p.role is LoanPartyRole.lender]
    assert lender.amount == Decimal(4)
    assert [
        (i.due_persian_year, i.due_persian_month, i.due_day_of_month, i.status)
        for i in lender.installments
    ] == [
        (1403, 1, 1, InstallmentStatus.paid),
        (1403, 3, 1, InstallmentStatus.unpaid),
    ]
    # The layout has no day concept — day=1 must NOT be reported as an issue.
    assert not [i for i in parsed_1403.issues if i.category is IssueCategory.missing_day]


# --------------------------------------------------------------------- fallbacks


def test_blank_loan_number_is_synthesised_with_warning() -> None:
    ws = _build_1401_sheet()
    ws.cell(row=7, column=2).value = None  # wipe loan 205's number
    result = ParseResult()
    parse_year(ws, 1401, LAYOUTS[1401], result)
    synth = result.loans[1]
    assert synth.loan_number == "بدون‌شماره-r7"
    issues = [i for i in result.issues if i.category is IssueCategory.orphan_row]
    assert issues
    assert issues[0].severity is IssueSeverity.warning


def test_missing_total_falls_back_to_lender_sum() -> None:
    ws = _build_1401_sheet()
    ws.cell(row=7, column=5).value = None  # wipe loan 205's total
    result = ParseResult()
    parse_year(ws, 1401, LAYOUTS[1401], result)
    loan = result.loans[1]
    assert loan.total_amount == Decimal(5)  # Σ lender amounts
    assert loan.parties[0].amount == Decimal(5)  # borrower party follows
    assert [i for i in result.issues if i.category is IssueCategory.total_mismatch]


def test_unfundable_loan_is_skipped_with_error() -> None:
    """No total AND no lender money → cannot satisfy total_amount > 0."""
    ws = _ws("سال 1401")
    ws.cell(row=3, column=1, value=1)
    ws.cell(row=3, column=2, value=204)
    ws.cell(row=3, column=6, value="قرض‌گیرنده")
    result = ParseResult()
    parse_year(ws, 1401, LAYOUTS[1401], result)
    assert result.loans == []
    errors = [i for i in result.issues if i.severity is IssueSeverity.error]
    assert errors
    assert errors[0].category is IssueCategory.orphan_row


def test_missing_borrower_keeps_loan_without_borrower_party() -> None:
    ws = _build_1401_sheet()
    ws.cell(row=7, column=6).value = None  # wipe loan 205's borrower
    result = ParseResult()
    parse_year(ws, 1401, LAYOUTS[1401], result)
    loan = result.loans[1]
    assert all(p.role is LoanPartyRole.lender for p in loan.parties)
    assert [i for i in result.issues if i.category is IssueCategory.unresolved_person]


def test_zero_amount_grid_cells_are_skipped() -> None:
    ws = _build_1403_sheet()
    ws.cell(row=5, column=14, value=0)  # zero amount cell (would violate CHECK)
    result = ParseResult()
    parse_year(ws, 1403, LAYOUTS[1403], result)
    (loan,) = result.loans
    (lender,) = [p for p in loan.parties if p.role is LoanPartyRole.lender]
    assert len(lender.installments) == 2  # the zero cell did not land
